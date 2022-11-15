#!/usr/bin/env python3

import tkinter as tk
import tkinter.ttk as ttk
from appdirs import user_config_dir
from pathlib import Path
from os.path import join as path_join
import configparser
from os.path import exists as path_exists
from os.path import isfile as file_exists
from tkinter import filedialog as fd
from tkinter import Canvas
from tkinter.messagebox import showerror, showinfo, showwarning
from shutil import move as move_file
from shutil import copy as copy_file
from numpy import imag
from openpyxl import load_workbook
from fillpdf import fillpdfs
from num2words import num2words
from math import trunc
from sys import platform
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, date, timedelta
import pandas as pd
import time
import math

def empty_validation(input):
    if not input or input == '' or input == None:
        return 0
    else:
        return 1

def empty_message_error():
    showerror(
        title='Error de validación',
        message='Los campos no pueden estar vacíos.'
    )

def separar_numeros(texto):
    numeros = False
    for letra in texto:
        if letra.isdigit():
            numeros = True
            break
    if numeros:
        for i, letra in enumerate(texto):
            if letra.isdigit():
                return [texto[:i], texto[i:]]
    return [texto]

def separar_ci(depositante):
    if 'C.I.' in depositante:
        datos = depositante.split('C.I.')
    elif 'CI:' in depositante:
        datos = depositante.split('CI:')
    elif 'CI.' in depositante:
        datos = depositante.split('CI.')
    elif 'C.I' in depositante:
        datos = depositante.split('C.I')
    elif 'CI,' in depositante:
        datos = depositante.split('CI,')
    elif 'C,I,' in depositante:
        datos = depositante.split('C,I,')
    elif '-' in depositante:
        datos = depositante.split('-')
    else:
        datos = depositante.split(' CI ')
    if len(datos) == 1:
        datos = separar_numeros(datos[0])
    for i in range(len(datos)):
        datos[i] = datos[i].strip('-').strip(':').strip(',').strip('.').strip()
    return datos

# UNINET
url = 'https://uninetplus.bancounion.com.bo/Uninetplus/Account/Login'
options = Options()
options.set_preference('profile', os.path.join(os.getcwd(), 'times.json'))
options.set_preference('intl.accept_languages', 'en-US, en')
options.set_preference('dom.push.enabled', False)
options.set_preference('dom.webnotifications.enabled', False)

# APP
app_name = 'IPSPAportes'
author_name = 'Daniel_Jimenez'
file_config_name = 'ipsp_aportes.ini'
file_pdf_name = 'PLANTILLA_FORMULARIO.pdf'
file_empty_message = 'Debe seleccionar un archivo'
path_empty_message = 'Debe seleccionar un directorio'
app_config_path = user_config_dir(app_name, author_name)
Path(app_config_path).mkdir(parents=True, exist_ok=True)
file_config_path = path_join(app_config_path, file_config_name)
file_pdf_path = path_join(app_config_path, file_pdf_name)

config = configparser.ConfigParser()
if path_exists(file_config_path):
    config.read(file_config_path)
if not config.has_section('INPUT'):
    config.add_section('INPUT')
if not config.has_section('SIGNER'):
    config.add_section('SIGNER')
if not config.has_section('UNINET'):
    config.add_section('UNINET')

root = tk.Tk()
s = ttk.Style(root)
s.theme_use('default')
s.configure('TNotebook.Tab', font=('Arial', '12', 'normal'))
s.configure('custom_button.TButton', font=('Arial', '12', 'normal'))

root.title('MAS IPSP - APORTES')
empty_validation_command = root.register(empty_validation)
empty_message_error_command = root.register(empty_message_error)

notebook = ttk.Notebook(root)
notebook.pack(fill=tk.BOTH, expand=True)
notebook.pressed_index = None

tab1 = ttk.Frame(notebook)
tab1.pack(fill=tk.BOTH, expand=True)

tab2 = ttk.Frame(notebook)
tab2.pack(fill=tk.BOTH, expand=True)

tab3 = ttk.Frame(notebook)
tab3.pack(fill=tk.BOTH, expand=True)

img1 = tk.PhotoImage(data='iVBORw0KGgoAAAANSUhEUgAAABAAAAAQCAMAAAAoLQ9TAAAABGdBTUEAALGPC/xhBQAAACBjSFJNAAB6JgAAgIQAAPoAAACA6AAAdTAAAOpgAAA6mAAAF3CculE8AAABv1BMVEUAAABGRkEqKiciIiAkJCJISEMODg4LCwtISEQFBQUCAgIDAwMDAwMDAwMDAwMDAwIDAwMCAgICAgIgIB4DAwMDAwMODg0DAwMLCwoPDw4ICAgCAgIBAQEBAQEHBwYDAwMEBAQSEhEBAQEBAQEHBwYMDAsEBAQEBAQEBAQGBgUQEBACAgEBAQECAgIICAcJCQkCAgICAgEFBQUGBgUGBgUHBwYFBQUBAQECAgEDAwNbW1UEBAMCAgIBAQEBAQEBAQEGBgYGBgYGBgYJCQgCAgICAgIKCgl4eHEDAwMCAgICAgICAgICAgIFBQQFBQUDAwMBAQFAQD1HR0NEREAEBAQVFRQNDQwQEA8SEhEYGBcJCQkBAQEDAwNVVVBEREAEBAMEBAQDAwMDAwMCAgIDAwM1NTINDQ0ODg0BAQEBAQEEBAQ2NjICAgIDAwMDAwMQEA8qKig3NzMdHRsJCQg2NjMEBAMEBAQDAwMEBAQiIiBhYVsCAgIDAwICAgIDAwJpaWNKSkVUVE+KioL///8KCgoCAgJTU08CAgIDAwMDAwMDAwMiIiFYWFMPDw8GBgYGBgUGBgUICAhLS0hKSkb///8SAPvFAAAAlHRSTlMAAwYHBgMMFQMtb2RiYWVrVY+ZB0ZRE2AZEh+B8b0jRUQMxtwcFjw9Pi8QsuiCHhykszEtLiQvybBLAkym5NjLKi0pFKCSFwJOhKOdmDkyXPYDAwRGDRURDwoQ4EwDA0k/VF+MXgUUFLfTNAWvTWMQBgUJHgVKQFZCCANwaZdnAwQDAgEZcwN5WFpZBQMNKCorGgIDCz5HMwAAAAFiS0dEg/y0z9IAAAAHdElNRQflDAwBKxxmyLk5AAAA8klEQVQY02NgZGKGARZWNnYOBk4ubh5ebj4g4OYXEBQSZhARFRMQlxCTBAIpaRlZOQZ5BUUlZRVVNXUNTS1tHV09oIC+gaGRpLGxiamZuYWlFYO8tY2tnb2Do6aTs4urpZs7g7yHp5e3j4mvmp9/gHRgUDCDfEhoWJhYeERklGx0TGxwHIN8fIJAYmKigEZSQHJKKkggJDQtPT1dTDIjM4s3G65CIFEtJ9ctLxuqIi2/oDDcsaiYuwQkUFpWXl5RWSVWXVNbVw8SUGBoaGhsam5p5eEObwMKJLR3dAJBF59PRXdPMFCgt68fDCZMnDQ5DggAHkREcJfzK3gAAAAldEVYdGRhdGU6Y3JlYXRlADIwMjEtMTItMTJUMDE6NDM6MjUrMDA6MDCcBOh7AAAAJXRFWHRkYXRlOm1vZGlmeQAyMDIxLTEyLTEyVDAxOjQzOjI1KzAwOjAw7VlQxwAAACB0RVh0c29mdHdhcmUAaHR0cHM6Ly9pbWFnZW1hZ2ljay5vcme8zx2dAAAAGHRFWHRUaHVtYjo6RG9jdW1lbnQ6OlBhZ2VzADGn/7svAAAAGHRFWHRUaHVtYjo6SW1hZ2U6OkhlaWdodAAxOTJAXXFVAAAAF3RFWHRUaHVtYjo6SW1hZ2U6OldpZHRoADE5MtOsIQgAAAAZdEVYdFRodW1iOjpNaW1ldHlwZQBpbWFnZS9wbmc/slZOAAAAF3RFWHRUaHVtYjo6TVRpbWUAMTYzOTI3MzQwNYMms/UAAAAPdEVYdFRodW1iOjpTaXplADBCQpSiPuwAAABWdEVYdFRodW1iOjpVUkkAZmlsZTovLy9tbnRsb2cvZmF2aWNvbnMvMjAyMS0xMi0xMi8wODEyMTA5YmFjYzExYjAwMGRlM2ZmNWQ5NzUwNjhkNy5pY28ucG5ncXv8dQAAAABJRU5ErkJggg==')
notebook.add(tab1, text='Generación de Formularios', image=img1, compound='left')

img2 = tk.PhotoImage(data='iVBORw0KGgoAAAANSUhEUgAAABEAAAAOCAMAAAD+MweGAAAABGdBTUEAALGPC/xhBQAAACBjSFJNAAB6JgAAgIQAAPoAAACA6AAAdTAAAOpgAAA6mAAAF3CculE8AAABU1BMVEUAAABEREBcXFUoKCgkJCQqKiokJCQpKSlZWVJEREBWVlDh4cooKCcjIyMhISE7Ozs7OzshISEjIyMpKSj+/uRKSkUoKCcjIyMjIyMjIyMiIiIwMDBtbW0jIyMjIyMjIyNISEMzMzEkJCMjIyMjIyMjIyMjIyMiIiIdHR0jIyMkJCNNTUgsLCsjIyMjIyMjIyMjIyNMTEdLS0YjIyMjIyMmJiUjIyMjIyMjIyMjIyMjIyMjIyMqKikrKyojIyMjIyMlJSUjIyMjIyMjIyMjIyMjIyMjIyMpKShPT0lLS0cjIyMjIyMjIyMjIyMlJSUjIyMjIyMjIyMpKShOTkkjIyMjIyNPT0pFRUFAQD0jIyMjIyMlJSUjIyMjIyMjIyMjIyMjIyMjIyMnJydBQT4jIyMjIyMjIyMsLCtCQj4jIyMjIyNDQz8/PzwjIyNBQT6enp7////oyIiYAAAAb3RSTlMAAwIRXaJVDwIEAwEVXLv187ZYFAEDFmLA9Pz5+PO9YAMIToGGhIOCg4JNAw1rkI+RAwOnuCKqtK+ws60QD8reKc7Z09TY0BMDA8jby9co0dLWEgPa1QMDBJajKpigm5yfmhgEdZKODARbWgQDzwPwxnlgAAAAAWJLR0Rw2ABsdAAAAAd0SU1FB+YBFwAgIryDce8AAADVSURBVAjXY2AEAiYGZhZWNnYGDhCHgZGTi5uHl48/X0BQSFiEixMoIiomLiEpJS0jLSUrJ88jysigoKikrKKqpq6hrqmqoqylqMCgraOrp6esD0T6yvoGujqGDIxGOsYmpmbmphaWplbWNtpAcwxt7ewdHJ0cnF0cXN3cPYAinrZe3g4+vn7+AX6BjkHBWEUMbX2AukIcnEP9XP2DwhgZwiPEIqOiY2Id4uIdEhKTksMZgnVSUg3S9IBIL80gNSU9mCEjMyEWARKyshlyVJ1zEcBZLQ8ANPkzm26T+okAAAAldEVYdGRhdGU6Y3JlYXRlADIwMjItMDEtMjNUMDA6MzI6MjYrMDA6MDC0OI9kAAAAJXRFWHRkYXRlOm1vZGlmeQAyMDIyLTAxLTIzVDAwOjMyOjI2KzAwOjAwxWU32AAAAABJRU5ErkJggg==')
notebook.add(tab2, text='Búsqueda UNINET', image=img2, compound='left')

img3 = tk.PhotoImage(data='iVBORw0KGgoAAAANSUhEUgAAABAAAAAQCAQAAAC1+jfqAAAABGdBTUEAALGPC/xhBQAAACBjSFJNAAB6JgAAgIQAAPoAAACA6AAAdTAAAOpgAAA6mAAAF3CculE8AAAAAmJLR0QA/4ePzL8AAAAJcEhZcwAAEHMAABBzARg5fEAAAAAHdElNRQflDAwBKwabqkBDAAABMElEQVQoz33Rv0vUARzG8dfd9zhNJQK/XYg/sLZcTq0hp7DJRfwD3LIil6KmaHFwjRoTJ1enliKhRsE4wQZFDkFFMI8MakjIQb+fhsu7IfG9Ps/z4fnwJJpc0a/VkfAfeYl2C75ZcwuFppTgspcm3XBfSZeiQQ9s+940zTgRToUQMpmw7FpdLMi5KgFVValhbUi1nuXbPJcJC/okOjz0U/isV65u+OiXsOV6o9WcsOaZewYgE8IHxUanR2oeW7FtXDGvJpDqaBh6VZXcsenAPEPeCsee/Pv+tlUvLAlHajYLvtpHi1lDKrqNee+msswlfyzm5M2bapz/7Y1+nQ6NeOWLnUTYMuBYRZ/wWo+yDcOeeueHk3ou1aOkYte0PUvWjTqHQRPKPll21wWkZyvU+Qt4YldxF3VvXAAAACV0RVh0ZGF0ZTpjcmVhdGUAMjAyMS0xMi0xMlQwMTo0Mjo1MiswMDowMLyktNIAAAAldEVYdGRhdGU6bW9kaWZ5ADIwMjEtMTItMTJUMDE6NDI6NTIrMDA6MDDN+QxuAAAAIHRFWHRzb2Z0d2FyZQBodHRwczovL2ltYWdlbWFnaWNrLm9yZ7zPHZ0AAAAYdEVYdFRodW1iOjpEb2N1bWVudDo6UGFnZXMAMaf/uy8AAAAYdEVYdFRodW1iOjpJbWFnZTo6SGVpZ2h0ADE5MkBdcVUAAAAXdEVYdFRodW1iOjpJbWFnZTo6V2lkdGgAMTky06whCAAAABl0RVh0VGh1bWI6Ok1pbWV0eXBlAGltYWdlL3BuZz+yVk4AAAAXdEVYdFRodW1iOjpNVGltZQAxNjM5MjczMzcyV0ymFAAAAA90RVh0VGh1bWI6OlNpemUAMEJClKI+7AAAAFZ0RVh0VGh1bWI6OlVSSQBmaWxlOi8vL21udGxvZy9mYXZpY29ucy8yMDIxLTEyLTEyLzNjOTliMDYxNmI0OWFkMmNkODUyNjNmMTEyNTA5NmM1Lmljby5wbmcNsPFfAAAAAElFTkSuQmCC')
notebook.add(tab3, text='Configuración', image=img3, compound='left')

tab1.wb = None
tab1.ws = None

tab2.wb = None
tab2.ws = None

input_sheet = tk.StringVar(tab1, '')

def load_excel_data(tab):
    if config.has_option('INPUT', 'path'):
        if config['INPUT']['path'] != file_empty_message:
            if file_exists(config['INPUT']['path']):
                if tab == 1:
                    tab1.wb = load_workbook(config['INPUT']['path'])
                    if len(tab1.wb.sheetnames) > 0:
                        input_sheet.set(tab1.wb.sheetnames[0])
                        tab1.ws = tab1.wb[tab1.wb.sheetnames[0]]
                        return True
                else:
                    tab2.wb = load_workbook(config['INPUT']['path'])
                    if len(tab2.wb.sheetnames) > 0:
                        input_sheet.set(tab2.wb.sheetnames[0])
                        tab2.ws = tab2.wb[tab2.wb.sheetnames[0]]
                        return True
    return False

pdf_form_path = tk.StringVar(root, config['FORM']['path'] if config.has_option('FORM', 'path') else file_empty_message)
output_path = tk.StringVar(root, config['OUTPUT']['path'] if config.has_option('OUTPUT', 'path') else path_empty_message)
excel_input_path = tk.StringVar(root, config['INPUT']['path'] if config.has_option('INPUT', 'path') else file_empty_message)

tab1.row_from = tk.IntVar(tab1, 2)
tab1.row_to = tk.IntVar(tab1, 2)
tab1.progress_total = tk.IntVar(tab1, 0)
tab1.progress_current = tk.IntVar(tab1, 0)
tab1.progress_success = tk.IntVar(tab1, 0)
tab1.progress = tk.StringVar(tab1, '0/0')

tab2.row_from = tk.IntVar(tab2, 2)
tab2.row_to = tk.IntVar(tab2, 2)
tab2.date_from = tk.StringVar(tab2, (date.today().replace(day=1)).strftime('%d/%m/%Y'))
tab2.date_to = tk.StringVar(tab2, date.today().strftime('%d/%m/%Y'))
tab2.driver = None
tab2.progress = tk.StringVar(tab2, '0%')
tab2.step = tk.IntVar(tab2, 0)

def select_file(file_types, config_section):
    filename = fd.askopenfilename(
        title='Abrir archivo',
        initialdir=str(Path.home()),
        filetypes=file_types
    )
    if filename:
        showinfo(
            title='Archivo seleccionado',
            message=filename
        )
        if not config.has_section(config_section):
            config.add_section(config_section)
        if config_section == 'FORM':
            pdf_form_path.set(filename)
        if config_section == 'INPUT':
            excel_input_path.set(filename)
        return True
    else:
        return False

def select_folder():
    pathname = fd.askdirectory(
        title='Seleccionar carpeta',
        initialdir=str(Path.home())
    )
    if pathname:
        showinfo(
            title='Carpeta seleccionada',
            message=pathname
        )
        if not config.has_section('OUTPUT'):
            config.add_section('OUTPUT')
        config.set('OUTPUT', 'path', pathname)
        output_path.set(pathname)
        save_config(False)
        tab1.progress.set('0/0')
        tab1.progress_success.set(0)
        return True
    else:
        return False

def select_excel_input_file():
    if (select_file([('Excel', '*.xlsx'), ('Excel', '*.XLSX'), ('Excel', '*.xls'), ('Excel', '*.XLS')], 'INPUT') == True):
        config.set('INPUT', 'path', excel_input_path.get())
        save_config(False)
        tab1.progress.set('0/0')
        tab1.progress_success.set(0)

def checkbox(data):
    return 'Off' if str(data) == 'NO' or str(data) == '' or data == None else 'SI'

def fill_pdf_template(data):
    data_dict = fillpdfs.get_form_fields(Path(pdf_form_path.get()))
    number = data[0]
    try:
        number = f'{int(number):05}'
    except:
        number = number if number != None else ''
    title = data[1] if data[1] != None else 'EFECTIVO'
    try:
        title = str(title).upper()
    except:
        title = 'EFECTIVO'
    deposit = data[4]
    if isinstance(deposit, str):
        deposit = deposit.strip()
    try:
        deposit = int(deposit)
    except:
        deposit = deposit if deposit != None else ''
    date_issue = data[2]
    try:
        date_issue = date_issue.strftime('%d/%m/%Y')
    except:
        date_issue = date_issue if date_issue != None else ''
    date_deposit = data[3]
    try:
        date_deposit = date_deposit.strftime('%d/%m/%Y')
    except:
        date_deposit = date_deposit if date_deposit != None else ''
    money_float = data[5]
    try:
        money_float = format(data[5], '.2f')
    except:
        money_float = money_float if money_float != None else ''
    money_literal = data[5]
    try:
        money_literal = '{0}{1}/100 BOLIVIANOS'.format(num2words(int(data[5])*100, lang='es', to='currency').split('euros')[0].split('euro')[0].upper(), f'{int(round(float(data[5])%1, 2)*100):02}')
    except:
        money_literal = money_literal if money_literal != None else ''
    try:
        month_year = str(data[6] if data[6] != None else '').upper()
        if data[7] != None and data[7] != '':
            month_year = month_year + ' DE '
            try:
                month_year = month_year + str(int(data[7])).upper()
            except:
                month_year = month_year + data[7]
    except:
        month_year = data[6] + ' DE ' + data[7]
    fill = {
        'number': number,
        'title': title,
        'attachment': '4' if title == 'EFECTIVO' else '5',
        'date_issue': date_issue,
        'date_deposit': date_deposit,
        'deposit': deposit,
        'money_float': money_float,
        'money_literal': money_literal,
        'month_year': month_year,
        'name': str(data[8] if data[8] != None else '').upper(),
        'ci': (trunc(int(data[9])) if isinstance(data[9], float) else str(data[9]).upper()) if data[9] != None else '',
        'tel': (trunc(int(data[10])) if isinstance(data[10], float) else str(data[10]).upper()) if data[10] != None else '',
        'check_militant': checkbox(data[11]),
        'check_monthly': checkbox(data[12]),
        'check_extraordinary': checkbox(data[13]),
        'check_generals': checkbox(data[14]),
        'check_municipal': checkbox(data[15]),
        'check_government': checkbox(data[16]),
        'check_education': checkbox(data[17]),
        'check_administrative': checkbox(data[18]),
        'check_others': checkbox(data[19])
    }
    for i in [1, 2]:
        data_dict['number{}'.format(i)] = fill['number']
        data_dict['title{}'.format(i)] = fill['title']
        data_dict['attachment{}'.format(i)] = fill['attachment']
        data_dict['date_issue{}'.format(i)] = fill['date_issue']
        data_dict['date_deposit{}'.format(i)] = fill['date_deposit']
        data_dict['money_float{}'.format(i)] = fill['money_float']
        data_dict['money_literal{}'.format(i)] = fill['money_literal']
        data_dict['month_year{}'.format(i)] = fill['month_year']
        data_dict['name{}'.format(i)] = fill['name']
        data_dict['ci{}'.format(i)] = fill['ci']
        data_dict['tel{}'.format(i)] = fill['tel']
        data_dict['check_militant{}'.format(i)] = fill['check_militant']
        data_dict['check_monthly{}'.format(i)] = fill['check_monthly']
        data_dict['check_extraordinary{}'.format(i)] = fill['check_extraordinary']
        data_dict['check_generals{}'.format(i)] = fill['check_generals']
        data_dict['check_municipals{}'.format(i)] = fill['check_municipal']
        data_dict['check_government{}'.format(i)] = fill['check_government']
        data_dict['check_education{}'.format(i)] = fill['check_education']
        data_dict['check_administrative{}'.format(i)] = fill['check_administrative']
        data_dict['check_others{}'.format(i)] = fill['check_others']
        data_dict['issuer_name{}'.format(i)] = config['SIGNER']['name'].upper()
        data_dict['issuer_charge{}'.format(i)] = config['SIGNER']['charge'].upper()
    out_file = path_join(Path(config['OUTPUT']['path']), '{}_{}.pdf'.format(fill['number'], fill['name'].replace(' ', '_')))
    try:
        if os.path.exists(out_file):
            os.remove(out_file)
        fillpdfs.write_fillable_pdf(Path(pdf_form_path.get()), out_file, data_dict, flatten=False)
        return True
    except:
        return False

def set_loading(loading):
    button_file['state'] = 'normal' if loading else 'disabled'
    button_folder['state'] = 'normal' if loading else 'disabled'
    entry_row_from['state'] = 'normal' if loading else 'disabled'
    entry_row_to['state'] = 'normal' if loading else 'disabled'
    uninet_row_from['state'] = 'normal' if loading else 'disabled'
    uninet_row_to['state'] = 'normal' if loading else 'disabled'
    button_run['state'] = 'normal' if loading else 'disabled'
    button_file_uninet['state'] = 'normal' if loading else 'disabled'
    button_navigator['state'] = 'normal' if loading else 'disabled'

def open_folder(folder):
    try:
        if platform.startswith('win32'):
            os.system('start '+r'{}'.format(folder))
        elif platform.startswith('linux'):
            os.system('xdg-open "{}"'.format(folder))
    except:
        return None

def generate_pdfs():
    if not file_exists(config['FORM']['path']):
        pdf_form_path.set(file_empty_message)
        showerror(
            title='Error en la configuración',
            message='El archivo plantilla PDF no existe, seleccione nuevamente.'
        )
        notebook.select(1)
        return None
    if not path_exists(config['OUTPUT']['path']):
        output_path.set(path_empty_message)
        showerror(
            title='Error en el carpeta de destino',
            message='La carpeta de destino seleccionado no existe, seleccione nuevamente.'
        )
        return None
    try:
        row_from = tab1.row_from.get()
        row_to = tab1.row_to.get()
    except:
        showerror(
            title='Error en la selección de filas',
            message='El valor de la selección de filas debe ser un número de entre las filas de la Hoja de Excel.'
        )
        return None
    if row_to < 2 or row_from < 2 or row_to > 100000 or row_from > 100000 or type(row_to) is not int or type(row_from) is not int:
        showerror(
            title='Error en la selección de filas',
            message='El valor de la selección de filas debe estar entre 2 y 100000.'
        )
        return None
    if row_to < row_from:
        showerror(
            title='Error en la selección de filas',
            message='El valor Hasta fila debe ser mayor o igual que el valor Desde fila.'
        )
        return None
    if not load_excel_data(1):
        excel_input_path.set(file_empty_message)
        showerror(
            title='Error al cargar la hoja de cálculo Excel',
            message='El archivo Excel no existe o no corresponde a la plantilla de datos.'
        )
        return None
    tab1.progress_success.set(0)
    tab1.progress_current.set(0)
    tab1.progress_total.set(row_to - row_from + 1)
    set_loading(False)
    for row in tab1.ws.iter_rows(min_row=row_from, max_col=20, max_row=row_to, values_only=True):
        tab1.progress_current.set(tab1.progress_current.get() + 1)
        tab1.progress.set('{0}/{1}'.format(tab1.progress_current.get(), tab1.progress_total.get()))
        escape = True
        for cell in row:
            if cell != None and cell != '':
                escape = False
                break
        if escape:
            continue
        else:
            result = fill_pdf_template(row)
            if result:
                tab1.progress_success.set(tab1.progress_success.get() + 1)
            else:
                showerror(
                    title='Error de almacenamiento',
                    message='Cierre los archivos PDF de aportes abiertos y vuelva a ejecutar el proceso.'
                )
                set_loading(True)
                tab1.update()
                return None
        tab1.update()
    set_loading(True)
    showinfo(
        title='Formularios generados',
        message='El proceso ha finalizado correctamente.'
    )
    if tab1.progress_success.get() > 1:
        open_folder(config['OUTPUT']['PATH'])

# Ventana generación de formularios

ttk.Label(tab1, text='Hoja de cálculo Excel de aportantes', font=('Arial', '14', 'bold underline')).grid(sticky='W', column=0, row=0, padx=10, pady=10, columnspan=4)

ttk.Label(tab1, text='Archivo Excel:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=1, padx=10, pady=5)
ttk.Label(tab1, text='', font=('Arial', '12', 'normal'), anchor='w', textvariable=excel_input_path, wraplength=340).grid(sticky='WE', column=1, row=1, padx=10, pady=5, columnspan=2)
button_file = ttk.Button(tab1, text='Seleccionar archivo', style='custom_button.TButton', command=select_excel_input_file)
button_file.grid(sticky='E', column=3, row=1, padx=10, pady=5)

ttk.Label(tab1, text='Carpeta donde se exportarán los formularios', font=('Arial', '14', 'bold underline')).grid(sticky='W', column=0, row=2, padx=10, pady=10, columnspan=4)

ttk.Label(tab1, text='Carpeta destino:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=3, padx=10, pady=5)
ttk.Label(tab1, text='', font=('Arial', '12', 'normal'), anchor='w', textvariable=output_path, wraplength=340).grid(sticky='WE', column=1, row=3, padx=10, pady=5, columnspan=2)
button_folder = ttk.Button(tab1, text='Seleccionar carpeta', style='custom_button.TButton', command=lambda:select_folder())
button_folder.grid(sticky='E', column=3, row=3, padx=10, pady=5)

ttk.Label(tab1, text='Desde fila:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=4, padx=10, pady=5)
entry_row_from = ttk.Entry(tab1, font=('Arial', '12', 'normal'), takefocus=0, textvariable=tab1.row_from)
entry_row_from.grid(sticky='WE', column=1, row=4, padx=10, pady=5)
ttk.Label(tab1, text='Hasta fila:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=2, row=4, padx=10, pady=5)
entry_row_to = ttk.Entry(tab1, font=('Arial', '12', 'normal'), takefocus=0, textvariable=tab1.row_to)
entry_row_to.grid(sticky='WE', column=3, row=4, padx=10, pady=5)

ttk.Label(tab1, text='Progreso:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=5, padx=10, pady=5)
ttk.Label(tab1, text='', font=('Arial', '12', 'normal'), anchor='w', textvariable=tab1.progress).grid(sticky='WE', column=1, row=5, padx=10, pady=5)
ttk.Label(tab1, text='Recibos generados:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=2, row=5, padx=10, pady=5)
ttk.Label(tab1, text='', font=('Arial', '12', 'normal'), anchor='w', textvariable=tab1.progress_success).grid(sticky='WE', column=3, row=5, padx=10, pady=5)

button_run = ttk.Button(tab1, text='Generar formularios', style='custom_button.TButton', command=generate_pdfs)
button_run.grid(sticky='E', column=3, row=6, padx=10, pady=20)

# Ventana UNINET

ttk.Label(tab2, text='Hoja de cálculo Excel de aportantes', font=('Arial', '14', 'bold underline')).grid(sticky='W', column=0, row=0, padx=10, pady=10, columnspan=4)

ttk.Label(tab2, text='Archivo Excel:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=1, padx=10, pady=5)
ttk.Label(tab2, text='', font=('Arial', '12', 'normal'), anchor='w', textvariable=excel_input_path, wraplength=340).grid(sticky='WE', column=1, row=1, padx=10, pady=5, columnspan=2)
button_file_uninet = ttk.Button(tab2, text='Seleccionar archivo', style='custom_button.TButton', command=select_excel_input_file)
button_file_uninet.grid(sticky='E', column=3, row=1, padx=10, pady=5)

ttk.Label(tab2, text='Desde fila:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=2, padx=10, pady=5)
uninet_row_from = ttk.Entry(tab2, font=('Arial', '12', 'normal'), takefocus=0, textvariable=tab2.row_from)
uninet_row_from.grid(sticky='WE', column=1, row=2, padx=10, pady=5)
ttk.Label(tab2, text='Hasta fila:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=2, row=2, padx=10, pady=5)
uninet_row_to = ttk.Entry(tab2, font=('Arial', '12', 'normal'), takefocus=0, textvariable=tab2.row_to)
uninet_row_to.grid(sticky='WE', column=3, row=2, padx=10, pady=5)

ttk.Label(tab2, text='Desde fecha:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=3, padx=10, pady=5)
uninet_date_from = ttk.Entry(tab2, font=('Arial', '12', 'normal'), takefocus=0, textvariable=tab2.date_from)
uninet_date_from.grid(sticky='WE', column=1, row=3, padx=10, pady=5)
ttk.Label(tab2, text='Hasta fecha:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=2, row=3, padx=10, pady=5)
uninet_date_to = ttk.Entry(tab2, font=('Arial', '12', 'normal'), takefocus=0, textvariable=tab2.date_to)
uninet_date_to.grid(sticky='WE', column=3, row=3, padx=10, pady=5)

def draw_captcha_image():
    tab2.driver.get(url)
    captcha_base64 = tab2.driver.find_element(By.XPATH, '//*[@title="Captcha Code"]').get_attribute('src')
    cabecera, captcha_base64 = captcha_base64.split(',', 1)
    captcha_image = tk.PhotoImage(data=captcha_base64)
    canvas.itemconfigure(image_container, image=captcha_image, state='normal')
    canvas.update()
    uninet_captcha_input['state'] = 'normal'
    button_captcha['state'] = 'normal'
    raise Exception('Imagen actualizada')

tab2.deposits = []
accounts_total = 1

def remaining_deposits():
    return len(list(filter(lambda d: d['success'] == False, tab2.deposits))) > 0

def open_navigator():
    tab2.progress.set('0%')
    tab2.step.set(0)
    if config['UNINET']['firefox'] == None or config['UNINET']['firefox'] == '':
        showerror(
            title='Ruta de firefox inexistente',
            message='Configure nuevamente la ruta de instalación de Firefox.'
        )
        notebook.select(2)
        return None
    else:
        options.binary_location = config['UNINET']['firefox']
    if config['UNINET']['user'] == None or config['UNINET']['user'] == '' or config['UNINET']['pass'] == None or config['UNINET']['pass'] == '':
        showerror(
            title='Credenciales de UNINET incorrectas',
            message='Configure nuevamente las credenciales de acceso a UNINET.'
        )
        notebook.select(2)
        return None
    if not file_exists(config['INPUT']['path']):
        config.set('INPUT', 'path', '')
        save_config(False)
        excel_input_path.set(file_empty_message)
        showerror(
            title='Error al cargar la hoja de cálculo Excel',
            message='El archivo Excel no existe o no corresponde a la plantilla de datos.'
        )
        return None
    try:
        row_from = tab2.row_from.get()
        row_to = tab2.row_to.get()
    except:
        showerror(
            title='Error en la selección de filas',
            message='El valor de la selección de filas debe ser un número de entre las filas de la Hoja de Excel.'
        )
        return None
    if row_to < 2 or row_from < 2 or row_to > 100000 or row_from > 100000 or type(row_to) is not int or type(row_from) is not int:
        showerror(
            title='Error en la selección de filas',
            message='El valor de la selección de filas debe estar entre 2 y 100000.'
        )
        return None
    if row_to < row_from:
        showerror(
            title='Error en la selección de filas',
            message='El valor Hasta fila debe ser mayor o igual que el valor Desde fila.'
        )
        return None
    try:
        date_from = tab2.date_from.get()
        date_to = tab2.date_to.get()
        dates = [date_from.split('/'), date_to.split('/')]
        for date in dates:
            if len(date) != 3:
                showerror(
                    title='Error en el formato de fecha',
                    message='El formato de fecha debe coincidir con: dd/mm/aaaa.'
                )
                return None
            if len(date[0]) != 2 or len(date[1]) != 2 or len(date[2]) != 4:
                showerror(
                    title='Error en el formato de fecha',
                    message='El formato de fecha debe coincidir con: dd/mm/aaaa.'
                )
                return None
        try:
            if datetime.strptime(date_to, '%d/%m/%Y') < datetime.strptime(date_from, '%d/%m/%Y'):
                showerror(
                    title='Error en las fechas',
                    message='La fecha Hasta debe ser mayor o igual que la fecha Desde.'
                )
                return None
        except:
            showerror(
                title='Error en el formato de fecha',
                message='El formato de fecha debe coincidir con: dd/mm/aaaa.'
            )
            return None
    except:
        showerror(
            title='Error en el formato de fecha',
            message='El formato de fecha debe coincidir con: dd/mm/aaaa.'
        )
        return None
    try:
        os.rename(config['INPUT']['path'], config['INPUT']['path'])
        tab2.deposits = []
        load_excel_data(2)
        for row in (range(tab2.row_from.get(), tab2.row_to.get() + 1)):
            deposit = tab2.ws.cell(row=row, column=5).value
            success = True
            if tab2.ws.cell(row=row, column=5).value != None and tab2.ws.cell(row=row, column=5).value != '':
                success = False
            if tab2.ws.cell(row=row, column=4).value != None and tab2.ws.cell(row=row, column=6).value != None and tab2.ws.cell(row=row, column=9).value != None and tab2.ws.cell(row=row, column=4).value != '' and tab2.ws.cell(row=row, column=6).value != '' and tab2.ws.cell(row=row, column=9).value != '':
                success = True
            try:
                deposit = int(deposit)
            except:
                deposit = None if deposit == '' else deposit
            tab2.deposits.append({
                'number': deposit,
                'row': row,
                'success': success
            })
    except:
        showerror(
            title='Error al abrir archivo Excel',
            message='Debe cerrar el archivo Excel para poder modificarlo.'
        )
        return None
    tab2.progressbar['maximum'] = len(tab2.deposits)
    tab2.update()
    if not remaining_deposits():
        showinfo(
            title='Datos obtenidos',
            message='El proceso ha finalizado correctamente.'
        )
        tab2.progress.set('100%')
        tab2.step.set(tab2.progressbar['maximum'])
        return None
    button_file_uninet['state'] = 'disabled'
    button_navigator['state'] = 'disabled'
    uninet_row_from['state'] = 'disabled'
    uninet_row_to['state'] = 'disabled'
    uninet_date_from['state'] = 'disabled'
    uninet_date_to['state'] = 'disabled'
    service = Service(os.path.join(os.getcwd(), 'geckodriver.exe' if os.name == 'nt' else 'geckodriver'))
    tab2.driver = webdriver.Firefox(options=options, service=service)
    draw_captcha_image()

def fill_login():
    if uninet_captcha_input.get() == None or uninet_captcha_input.get() == '':
        showwarning(
            title='Texto captcha faltante',
            message='Debe llenar el texto captcha de la imagen.'
        )
        return None
    uninet_captcha_input['state'] = 'disabled'
    button_captcha['state'] = 'disabled'
    input_user = tab2.driver.find_element(By.ID, 'Usuario')
    input_user.clear()
    input_user.send_keys(config['UNINET']['user'])
    tab2.driver.find_element(By.ID, 'Captcha').send_keys(uninet_captcha_input.get())
    tab2.driver.find_element(By.XPATH, '//input[@value="Login"]').click()
    try:
        WebDriverWait(tab2.driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/section/form/div[3]/div[1]/div[2]/div[7]')))
    except:
        pass
    error = tab2.driver.find_elements(By.XPATH, '/html/body/div[2]/section/form/div[3]/div[1]/div[2]/div[7]')
    if len(error) > 0:
        if error[0].text == 'The capcha code does not correspond with the image.':
            uninet_captcha.set('')
            draw_captcha_image()
    else:
        try:
            WebDriverWait(tab2.driver, 5).until(EC.presence_of_element_located((By.ID, 'VerificaAlias')))
            tab2.driver.find_element(By.ID, 'VerificaAlias').click()
        except:
            print('No se encontró La imágen de verificación')
        tab2.driver.find_element(By.ID, 'Password').send_keys(config['UNINET']['pass'])
        tab2.driver.find_element(By.ID, 'btn-login').click()
        try:
            WebDriverWait(tab2.driver, 5).until(EC.presence_of_element_located((By.XPATH, '//div[contains(@class, "alert-box error") and text()="The password entered is not valid or your account (user) is inactive. Remember that after three consecutive attempts your account becomes inactive."]')))
            error = tab2.driver.find_elements(By.XPATH, '//div[contains(@class, "alert-box error") and text()="The password entered is not valid or your account (user) is inactive. Remember that after three consecutive attempts your account becomes inactive."]')
            if len(error) > 0:
                tab2.driver.quit()
                uninet_captcha.set('')
                config.set('UNINET', 'pass', '')
                save_config(False)
                uninet_pass.set('')
                showerror(
                    title='Credenciales de UNINET incorrectas',
                    message='Configure nuevamente las credenciales de acceso a UNINET.'
                )
                notebook.select(2)
                button_file_uninet['state'] = 'normal'
                button_navigator['state'] = 'normal'
                uninet_row_from['state'] = 'normal'
                uninet_row_to['state'] = 'normal'
                uninet_date_from['state'] = 'normal'
                uninet_date_to['state'] = 'normal'
                uninet_captcha_input['state'] = 'disabled'
                button_captcha['state'] = 'disabled'
                return None
        except:
            pass
        try:
            WebDriverWait(tab2.driver, 15).until(EC.presence_of_element_located((By.CLASS_NAME, 'modal-content')))
        except:
            print('No es necesario cerrar ningún modal')
        modal = tab2.driver.find_elements(By.CSS_SELECTOR, '#Notificacion > div:nth-child(1) > div:nth-child(1) > div:nth-child(3) > button:nth-child(1)')
        if len(modal) > 0:
            modal[0].click()
        accounts_total = goto_account()
        remaining = True
        for account in range(1, accounts_total):
            goto_account(account)
            time.sleep(2)
            WebDriverWait(tab2.driver, 15).until(EC.presence_of_element_located((By.ID, 'Extracto_5')))
            tab2.driver.find_element(By.ID, 'Extracto_5').click()
            WebDriverWait(tab2.driver, 15).until(EC.presence_of_element_located((By.ID, 'start')))
            WebDriverWait(tab2.driver, 15).until(EC.presence_of_element_located((By.ID, 'end')))
            tab2.driver.execute_script('arguments[0].setAttribute("value", "{}")'.format(tab2.date_from.get()), tab2.driver.find_element(By.ID, 'start'))
            tab2.driver.execute_script('arguments[0].setAttribute("value", "{}")'.format(tab2.date_to.get()), tab2.driver.find_element(By.ID, 'end'))
            WebDriverWait(tab2.driver, 15).until(EC.presence_of_element_located((By.ID, 'btn-continuar')))
            tab2.driver.find_element(By.ID, 'btn-continuar').click()
            try:
                WebDriverWait(tab2.driver, 5).until(EC.presence_of_element_located((By.XPATH, '//div[contains(@class, "alert-box error") and text()="No search results found"]')))
                error = tab2.driver.find_elements(By.XPATH, '//div[contains(@class, "alert-box error") and text()="No search results found"]')
                if len(error) > 0:
                    continue
            except:
                pass
            try:
                WebDriverWait(tab2.driver, 15).until(EC.presence_of_element_located((By.ID, 'CantidadMovimientos')))
                amount_select = Select(tab2.driver.find_element(By.ID, 'CantidadMovimientos'))
                amount_select.select_by_index(1)
            except:
                pass
            # Paginación
            try:
                WebDriverWait(tab2.driver, 15).until(EC.presence_of_element_located((By.CLASS_NAME, 'pagination')))
                paginacion = tab2.driver.find_element(By.CLASS_NAME, 'pagination')
                total_pages = len(paginacion.find_elements(By.TAG_NAME, 'li'))
                last_li = tab2.driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div/div[2]/div[2]/form/div[2]/div/div[2]/ul/li[{}]/a'.format(total_pages))
                total_pages = int(last_li.text)
            except:
                total_pages = 1
            for page in range(1, total_pages+1):
                time.sleep(2)
                try:
                    WebDriverWait(tab2.driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'cambio-extractos') and text()='{}']".format(page))))
                    tab2.driver.find_element(By.XPATH, "//a[contains(@class, 'cambio-extractos') and text()='{}']".format(page)).click()
                except:
                    pass
                time.sleep(4)
                WebDriverWait(tab2.driver, 15).until(EC.presence_of_element_located((By.ID, 'no-more-tables')))
                div = tab2.driver.find_element(By.ID, 'no-more-tables')
                table1 = div.find_elements(By.TAG_NAME, 'table')
                df1 = pd.read_html(table1[0].get_attribute('outerHTML'))[0]
                uninet_deposits = list(df1.get('No. Document.'))
                for deposit_index in range(0, len(tab2.deposits)):
                    deposit = tab2.deposits[deposit_index]
                    progress = int(math.ceil(100 * tab2.step.get() / tab2.progressbar['maximum']))
                    tab2.progress.set('{0}% - Cuenta: {1}/{2} - Página: {3}/{4} - Fila: {5} - Buscando documento: {6}'.format(progress, account, accounts_total-1, page, total_pages, deposit['row'], deposit['number'] if deposit['number'] != None else '-'))
                    tab2.update()
                    if deposit['number'] != None and deposit['success'] == False:
                        if deposit['number'] in uninet_deposits:
                            index = uninet_deposits.index(deposit['number']) + 1
                            date = df1.iloc[index-1]['Date']
                            try:
                                date = datetime.strptime(date, '%d/%m/%Y')
                            except:
                                pass
                            money = float(df1.iloc[index-1]['Amount'])
                            tab2.ws.cell(row=deposit['row'], column=4).value = date
                            tab2.ws.cell(row=deposit['row'], column=6).value = money
                            tab2.wb.save(config['INPUT']['path'])
                            # Buscar nombre y CI
                            description = df1.iloc[index-1]['Description'].upper().strip()
                            tag_visible = False
                            while not tag_visible:
                                try:
                                    WebDriverWait(tab2.driver, 15).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/div/div/div/div[2]/div[2]/form/div[2]/div/div[2]/div/table/tbody/tr[{}]/td[7]/center/button'.format(index))))
                                    elements = tab2.driver.find_elements(By.XPATH, '/html/body/div[2]/div[1]/div/div/div/div[2]/div[2]/form/div[2]/div/div[2]/div/table/tbody/tr[{}]/td[7]/center/button'.format(index))
                                    if len(elements) > 0:
                                        tag_visible = True
                                except:
                                    pass
                            tab2.driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div/div[2]/div[2]/form/div[2]/div/div[2]/div/table/tbody/tr[{}]/td[7]/center/button'.format(index)).click()
                            time.sleep(2)
                            # Modal detalle
                            WebDriverWait(tab2.driver, 15).until(EC.presence_of_element_located((By.CLASS_NAME, 'modal-dialog')))
                            div2 = tab2.driver.find_element(By.CLASS_NAME, 'modal-dialog')
                            div3 = div2.find_element(By.CLASS_NAME, 'modal-body')
                            table2 = []
                            while len(table2) == 0:
                                try:
                                    WebDriverWait(tab2.driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.modal.fade.in#DetalleMovimiento')))
                                    WebDriverWait(tab2.driver, 15).until(EC.presence_of_element_located((By.ID, 'DetalleMovimientoModalLabel')))
                                    ActionChains(tab2.driver).move_to_element(tab2.driver.find_element(By.ID, 'DetalleMovimientoModalLabel')).perform()
                                    table2 = div3.find_elements(By.TAG_NAME, 'table')
                                except:
                                    pass
                            table2 = table2[0].get_attribute('outerHTML')
                            df2 = pd.read_html(table2)[0]
                            if description == 'DEPOSITO A CUENTA':
                                name_ci = separar_ci(df2[1][2].upper().strip())
                            elif description == 'N/C POR TRASPASO ENTRE BANCOS ACH':
                                name_ci = separar_ci(df2[1][3].upper().strip())
                            elif 'N/C TRASP.' in description:
                                name_ci = df2[1][4].upper().strip()
                                name_ci = name_ci.split('-')
                                if len(name_ci) > 1:
                                    name_ci = separar_ci(name_ci[1].upper().strip())
                                else:
                                    continue
                            else:
                                continue
                            name = name_ci[0].strip()
                            ci = False
                            if len(name_ci) == 2:
                                ci = name_ci[1].strip()
                            tab2.ws.cell(row=deposit['row'], column=9).value = name
                            if ci != False:
                                tab2.ws.cell(row=deposit['row'], column=10).value = ci
                            tag_visible = False
                            while not tag_visible:
                                try:
                                    ActionChains(tab2.driver).move_to_element(tab2.driver.find_element(By.ID, 'page-wrapper')).click().perform()
                                except:
                                    pass
                                try:
                                    tab2.driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div/div[2]/div[2]/form/div[2]/div/div[4]/div/div/div/div[3]/button[2]').click()
                                except:
                                    pass
                                try:
                                    tab2.driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div/div[1]/div[2]/form/div[2]/div/div[4]/div/div/div/div[1]/button/span').click()
                                except:
                                    pass
                                try:
                                    WebDriverWait(tab2.driver, 3).until(EC.visibility_of_element_located((By.ID, 'DetalleMovimientoModalLabel')))
                                except:
                                    tag_visible = True
                            tab2.deposits[deposit_index]['success'] = True
                            tab2.wb.save(config['INPUT']['path'])
                            tab2.step.set(tab2.step.get() + 1)
                            tab2.update()
                    remaining = remaining_deposits()
                    if not remaining:
                        break
                if not remaining:
                    break
            if not remaining:
                break
    tab2.driver.quit()
    tab2.step.set(tab2.progressbar['maximum'])
    tab2.update()
    button_file_uninet['state'] = 'normal'
    button_navigator['state'] = 'normal'
    uninet_row_from['state'] = 'normal'
    uninet_row_to['state'] = 'normal'
    uninet_date_from['state'] = 'normal'
    uninet_date_to['state'] = 'normal'
    uninet_captcha_input['state'] = 'disabled'
    button_captcha['state'] = 'disabled'
    uninet_captcha.set('')
    canvas.itemconfigure(image_container, state='normal')
    canvas.update()
    showinfo(
        title='Datos obtenidos de UNINET',
        message='El proceso ha finalizado correctamente.'
    )
    tab2.progress.set('100%')
    tab2.update()
    raise Exception('Imagen actualizada')

def goto_account(account=None):
    WebDriverWait(tab2.driver, 15).until(EC.element_to_be_clickable((By.ID, 'link-1')))
    tab2.driver.execute_script('arguments[0].click();', tab2.driver.find_element(By.ID, 'link-1'))
    WebDriverWait(tab2.driver, 15).until(EC.element_to_be_clickable((By.ID, 'link-12')))
    tab2.driver.find_element(By.ID, 'link-12').click()
    WebDriverWait(tab2.driver, 15).until(EC.element_to_be_clickable((By.ID, 'CuentaOrigens_CuentaOrigen')))
    account_select = Select(tab2.driver.find_element(By.ID, 'CuentaOrigens_CuentaOrigen'))
    if account == None:
        return len(account_select.options)
    else:
        account_select.select_by_index(account)
        return None

button_navigator = ttk.Button(tab2, text='Buscar datos de depósitos', style='custom_button.TButton', command=open_navigator)
button_navigator.grid(sticky='E', column=0, row=4, padx=10, pady=20, columnspan=4)

ttk.Label(tab2, text='Ingrese el texto de la imágen:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=1, row=5, padx=10, pady=5)

canvas = Canvas(tab2, width=150, height=60)
canvas.grid(sticky='WE', column=0, row=5, padx=10, pady=5)
image_container = canvas.create_image(75, 30)

uninet_captcha = tk.StringVar(tab2, '')
uninet_captcha_input = ttk.Entry(tab2, font=('Arial', '12', 'normal'), takefocus=0, textvariable=uninet_captcha, validate='focusout', validatecommand=(empty_validation_command, '%P'), invalidcommand=(empty_message_error_command), state='disabled')
uninet_captcha_input.grid(sticky='WE', column=2, row=5, padx=10, pady=5)

button_captcha = ttk.Button(tab2, text='Ingresar', style='custom_button.TButton', state='disabled', command=fill_login)
button_captcha.grid(sticky='E', column=3, row=5, padx=10, pady=20)

ttk.Label(tab2, text='Progreso:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=6, padx=10, pady=5)
tab2.progressbar = ttk.Progressbar(tab2, maximum=100, variable=tab2.step)
tab2.progressbar.grid(sticky='WE', column=1, row=6, padx=10, pady=5, columnspan=3)

ttk.Label(tab2, textvariable=tab2.progress, font=('Arial', '12', 'normal'), anchor='e').grid(sticky='W', column=1, row=7, padx=10, pady=5, columnspan=3)

# Ventana configuración

ttk.Label(tab3, text='Datos de la autoridad firmante', font=('Arial', '14', 'bold underline')).grid(sticky='W', column=0, row=0, padx=10, pady=10, columnspan=4)

ttk.Label(tab3, text='Nombre:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=1, padx=0, pady=5)
signer_name = tk.StringVar(root, config['SIGNER']['name'] if config.has_option('SIGNER', 'name') else '')
ttk.Entry(tab3, font=('Arial', '12', 'normal'), takefocus=0, textvariable=signer_name, validate='focusout', validatecommand=(empty_validation_command, '%P'), invalidcommand=(empty_message_error_command)).grid(sticky='WE', column=1, row=1, padx=10, pady=5, columnspan=3)

ttk.Label(tab3, text='Cargo:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=2, padx=0, pady=5)
signer_charge = tk.StringVar(root, config['SIGNER']['charge'] if config.has_option('SIGNER', 'charge') else '')
ttk.Entry(tab3, font=('Arial', '12', 'normal'), takefocus=0, textvariable=signer_charge, validate='focusout', validatecommand=(empty_validation_command, '%P'), invalidcommand=(empty_message_error_command)).grid(sticky='WE', column=1, row=2, padx=10, pady=5, columnspan=3)

ttk.Label(tab3, text='Plantilla de formulario pdf', font=('Arial', '14', 'bold underline')).grid(sticky='W', column=0, row=3, padx=10, pady=10, columnspan=4)

ttk.Label(tab3, text='Formulario PDF:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=4, padx=10, pady=5)
ttk.Label(tab3, text='', font=('Arial', '12', 'normal'), anchor='w', textvariable=pdf_form_path, wraplength=340).grid(sticky='WE', column=1, row=4, padx=10, pady=5, columnspan=2)
ttk.Button(tab3, text='Seleccionar archivo', style='custom_button.TButton', command=lambda:select_file([('pdf file', '*.pdf'), ('pdf file', '*.PDF')], 'FORM')).grid(sticky='E', column=3, row=4, padx=10, pady=5)

ttk.Label(tab3, text='Datos de acceso a UNINET', font=('Arial', '14', 'bold underline')).grid(sticky='W', column=0, row=5, padx=10, pady=10, columnspan=4)

ttk.Label(tab3, text='Usuario:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=6, padx=0, pady=5)
uninet_user = tk.StringVar(root, config['UNINET']['user'] if config.has_option('UNINET', 'user') else '')
ttk.Entry(tab3, font=('Arial', '12', 'normal'), takefocus=0, textvariable=uninet_user, validate='focusout', validatecommand=(empty_validation_command, '%P'), invalidcommand=(empty_message_error_command)).grid(sticky='WE', column=1, row=6, padx=10, pady=5, columnspan=3)

ttk.Label(tab3, text='Contraseña:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=7, padx=0, pady=5)
uninet_pass = tk.StringVar(root, config['UNINET']['pass'] if config.has_option('UNINET', 'pass') else '')
ttk.Entry(tab3, font=('Arial', '12', 'normal'), takefocus=0, textvariable=uninet_pass, validate='focusout', validatecommand=(empty_validation_command, '%P'), invalidcommand=(empty_message_error_command), show="*").grid(sticky='WE', column=1, row=7, padx=10, pady=5, columnspan=3)

ttk.Label(tab3, text='Ruta Firefox:', font=('Arial', '12', 'normal'), anchor='e').grid(sticky='WE', column=0, row=8, padx=0, pady=5)
firefox_path = tk.StringVar(root, config['UNINET']['firefox'] if config.has_option('UNINET', 'firefox') else '')
ttk.Entry(tab3, font=('Arial', '12', 'normal'), takefocus=0, textvariable=firefox_path, validate='focusout', validatecommand=(empty_validation_command, '%P'), invalidcommand=(empty_message_error_command)).grid(sticky='WE', column=1, row=8, padx=10, pady=5, columnspan=3)

def save_config(show_info=True):
    signer_name_value = signer_name.get().strip().upper()
    signer_charge_value = signer_charge.get().strip().upper()
    uninet_user_value = uninet_user.get().strip()
    uninet_pass_value = uninet_pass.get().strip()
    firefox_path_value = firefox_path.get().strip()
    if not signer_name_value or not signer_charge_value or not uninet_user_value or not uninet_pass_value or not firefox_path_value:
        empty_message_error()
        notebook.select(1)
    else:
        config.set('SIGNER', 'name', signer_name_value)
        config.set('SIGNER', 'charge', signer_charge_value)
        config.set('UNINET', 'user', uninet_user_value)
        config.set('UNINET', 'pass', uninet_pass_value)
        config.set('UNINET', 'firefox', firefox_path_value)
        if file_pdf_path != pdf_form_path.get():
            copy_file(pdf_form_path.get(), file_pdf_path+'_TMP')
            move_file(file_pdf_path+'_TMP', file_pdf_path)
            config.set('FORM', 'path', file_pdf_path)
            pdf_form_path.set(file_pdf_path)
        with open(file_config_path, 'w') as configfile:
            config.write(configfile)
        if show_info == True:
            showinfo(title='Configuración', message='Configuración guardada exitosamente.')
            notebook.select(0)

ttk.Button(tab3, text='Guardar', style='custom_button.TButton', command=save_config).grid(sticky='E', column=3, row=9, padx=10, pady=20)

if path_exists(file_config_path):
    notebook.select(0)
else:
    notebook.select(2)

icon = tk.PhotoImage(data='iVBORw0KGgoAAAANSUhEUgAAACAAAAAgCAIAAAD8GO2jAAAABGdBTUEAALGPC/xhBQAAACBjSFJNAAB6JgAAgIQAAPoAAACA6AAAdTAAAOpgAAA6mAAAF3CculE8AAAJkklEQVRIx7VWaXRV1RXe59xzhzcm7+W9DI9MZIAEQogZIMFAkMhQsTK6arVLEahVsctqq6gtVGnFrmWXdFlLUavUsThgoQrKoFAhAUQkoUASkpAByPBeQkLuG+5w7jn9cQP+6d+eX/u7Z5+91t537+/bKFCzCf6fB/P/a3gAApwjjDBGwAEQMMYZ4wghASPbgwMwxgGAc45veAJYjHPOEUKcc4wxxogzzoEjhBBCzGKMc4QQwQhpOk1oJkKIce5UREUmjPGRMY1xDgACxm6nBAAYI02ncc3ECAGA1y3b0QWM45qR0CghGCHgHChlHrcsEswYJ2pcX1xXfP+yihE1keRW/vbRN/sbO5K9ypanbk9NcSOEBofU57Z+yTmPxo0FswpXr6hM6FSNahv/fFA3KCE4GtNLizLuX1ZRVhRyOaXLA9eON/e+/tFJNabJEiGGYU3KDSyeW6QZVJGIbtB/HThXXJb98N01jHGMUU/fyO+2fUUpFwT8qzVzastzDdOSRKHxdO/7e5oUiVSWZO55dZXbKTe39keuxuZV5y+snbRiQcn8NW8YhoUBgcUYAIyOJWIJo64qz+111JbnAkBv/ygAUIthhNSYPqcit7Y891pU6+0bAYDVKyodshjTjHnV+W6nbFLr77tOLVn31vo/fv7mzpMN3/UEfS6TWhg42H+zLzzW8F130O+qKcueXZkLAP8+eREAMEIIAWN81bJKADh9vm/za4cBoLYit7Yih1LWeLonoZkiEV5avzjSuOGuxdNHVO2l7Ufau4cciohv9JNJ2f6GDgB45J6a6unZTS19Z9sH7atYwiydnH773CIA2N9wYcfe5v6ICgD3La1QZPGro+3Tl/zpxTe/Pt8ZBkBVJZm/XDW7efejM0qzonGDwHjXgUMmR7/rNkx625wiQvD+xvaRsQQAIIQSmnnf0gqHIloWW3d3zb1LK7xuGQDumFc8JT+1smTCpNzg0Eis5q6tIhGeemDuE6vnKLI4f1bhkVPd5EaJHIrY1hVpbh2ompYJAAcbO/Iy/QCg6ebETN+PflAKACNjiWjckEVhIKJmh5Ilkdy5aJosCo+tmg0AeVn+9z9rSvW77LzPtg+IBH+fAcYonjAONLZXTcscHFJPnr1cUphm1+fHi8tSU9wA8Py2Qy+/25DsccgSOfreg3lZ/vuWVixY80Yo1bt8QcnalVVrV1YhhK6p2otvfv3Z4dYkj4J8M571eR2pAXdCM3v7Rl1OKSs9SY3p3VdGAz5nRtCT0CgRkCgShKCnb0SNGQDcpCw35EvxOQHg4qWr4eFoUV5qRtAjS0I0bnRfGRkcjrocEgCglOrnEEIYj3ODSZlti6KgG5ZlMc2gsajmcimGQTFGobQkt1NCAAPD0cHwmKyIyR7F5ZTimmkYlFpcJJhazDAsl1MCACBTn35g4864ZugGffDZTxb99A3DpH3hsdGxhGla19REZ+/wpq0HnWW/KV/+8r6jF8aiGufcsljkanT7J9+6yjf85Ikdcc3oC4+pMc2y2KiaaO8Z+vWWLzwVG4M1mzDjXJFFhyxKouCQRZEIIhEygp7BYfXDL86MRfW8LP+Gh+qX3jr1hccXLri5MDwcXfnou395/1jA51q1rOLJNXUGtRyymBH0XBq4tmNvcyxuFGSn/P4XCxfNnnQtqmGbJm8Q53UTPt1/7p47tjz7ygEbTi/KCPrdACASISsj+dCJzide3LvskXd27G32OGXbZ+ee5nvu2LJ52yEbFuYEKGUEYLxNx43rYPltpUWHnp5ZmmXDAw3tTS392zevzA4lv7R+MQAMjcSPN/c+vGmXSS3b564l5eXTs6qnZ9vw1LkrskTI/1QJxngo1etLcsQT5okzl1774MSxpl7doE0tfbdU58+YlllTllOYk/LDW4p9Xse7n562n2SmJQWSnbGEcayr9687jh851e12yQQB3BA1zscBxuj1j755/A97gn6XGtNVVZs3q2DrxqVEwK9+eGLVA9sd6UkNH6y7qTgUSvXaU40xeuW9xme27Av6XaNqwrK41y0zxggHwNfFC2OE0Pd5UItRizkUUZFIW1fEstjkicHNjy1aMKvQpGxKfioAfHq4pS+sjlcYjT9xOWSEwLIYQogQAUeuxs51DBIBD4/GYnGj9WIYIdQfURWZYIQsxgQBR2PGsp+/s37t3Oqy7JnTswGgtSuy+8vzL7x2eHHd5NaLYYxRf0R1yIQI2DAtW3cBAAVqNlkWsxizv0qioOkUIbAokxVREgWLccOgIhEMalHNJBJRFJFaTEuY3KRJfhe1eCymA+eyInJAnPMkj2J3JkKIcM4JwdzkkycGc0K+gSHV53UYplVSmHahe6i3fzTJreSEktW4oRvUpFbQ5xoejfuTnE5F1Ax6oXvI7ZSmFqRJotB1ZSTZo4gE//PgeUkUrm8VABihuGaWFWX8dt2tzW39kycGWzrDLZ3hDQ/Vv7XrVJJHmZjpRwgcsqjp1O2ULg9eK8wJnDhzyamI5VMndPQMtXVF6qryOnqHdYM6HZIg4Hd2n/YnOWw1RHb/6IYVHo6mJDs7eoYJwV638o89TbYCN7X02ZuHIKD+iEoEbFlM003dpG1dkbLi0Mf7zrZejOz9uk2WCOecMW4xxjjHGKFAzSabW9IDnvSgh1I2FtM9LmlOZd6Wt47UVxcMDKljUX1KQeromGaYVJYIALidks/rGFW1lovhUNA7PBqfmOk/1tRTX1NABLxz/9n0gNvecQQ9qXr5/JKN6+pbuyL1NQXbP/wmpptPrpkb8LkUmSyfP7WsOIQxunPBNJNatZW53569sn5t3ckzl6qmZX287+z8mwtnlGbfu7Q8GtNXr6hi1KopyymbErpzUemi2smnW/sET079Mz+bO7UgzaGIaQHP8f9cFkWhfmZ+SWFaXVXeweMdAPD4qtnvfdZ0vjN82+zJE9K886rzAz5XfnbK+c7Bm4onFOYETMoqp2YCwNyZ+U6HWFKY/smBc+kBN6VMKKxYkjvBd6ChPeBzuRzSpNzA0W+7Jmb6OcDbu0/Z8vn8tkMrFpZ4nHJb91BainvXwfP2GJZPCekGbe2MMMYHh6NNrf3tPUOtXZHPj7TNKstRY8bbu0+jQM1z0bhhMS4RAQAIwQLGJrUwQgndlETCOWecCxgzxhBCgMAwLCLg8T0VgFpMlohhUEAIOMcYMQ4YAWNckgjhHLxuBQEwzhEgex9VZBGAKzJhnAMghIBzjgABAAfudkic26Q1Pk2cc4nINzgfAXAYj/ZfT+fNPLpmHiMAAAAldEVYdGRhdGU6Y3JlYXRlADIwMjEtMTItMjdUMDg6NDk6MTMtMDU6MDApqNU+AAAAJXRFWHRkYXRlOm1vZGlmeQAyMDIxLTEyLTI3VDA4OjQ5OjEzLTA1OjAwWPVtggAAAABJRU5ErkJggg==')
root.tk.call('wm', 'iconphoto', root._w, icon)

root.geometry('')
root.resizable(False, False)
root.mainloop()
